"""
invoice_bot.py — Efraín Rojas Artavia
Extracts invoice data from PDFs, validates it, detects duplicates,
saves results to Excel/CSV and generates a summary report.
"""

import os
import re
import logging
from datetime import datetime
from pathlib import Path

import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fpdf import FPDF

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("invoice_bot.log"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

INPUT_DIR  = Path("invoices")
OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)
INPUT_DIR.mkdir(exist_ok=True)


# ── PDF Extraction ─────────────────────────────────────────────────────────────
def extract_invoice_data(pdf_path: Path) -> dict:
    """Extract key fields from an invoice PDF using regex patterns."""
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += (page.extract_text() or "") + "\n"

    def find(patterns, txt):
        for pat in patterns:
            m = re.search(pat, txt, re.IGNORECASE)
            if m:
                return m.group(1).strip()
        return ""

    invoice_number = find([
        r"invoice\s*#?\s*[:\-]?\s*([A-Z0-9\-]+)",
        r"factura\s*#?\s*[:\-]?\s*([A-Z0-9\-]+)",
        r"inv\s*[:\-]\s*([A-Z0-9\-]+)",
        r"n[°º]\s*[:\-]?\s*([A-Z0-9\-]+)",
    ], text)

    date = find([
        r"date\s*[:\-]\s*(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})",
        r"fecha\s*[:\-]\s*(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})",
        r"issued\s*[:\-]\s*(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4})",
        r"(\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{4})",
    ], text)

    vendor = find([
        r"from\s*[:\-]\s*(.+)",
        r"vendor\s*[:\-]\s*(.+)",
        r"proveedor\s*[:\-]\s*(.+)",
        r"bill\s*from\s*[:\-]\s*(.+)",
        r"company\s*[:\-]\s*(.+)",
    ], text)

    total = find([
        r"total\s*[:\-]\s*\$?\s*([\d,]+\.?\d*)",
        r"amount\s*due\s*[:\-]\s*\$?\s*([\d,]+\.?\d*)",
        r"grand\s*total\s*[:\-]\s*\$?\s*([\d,]+\.?\d*)",
        r"total\s*a\s*pagar\s*[:\-]\s*\$?\s*([\d,]+\.?\d*)",
        r"\$\s*([\d,]+\.\d{2})",
    ], text)

    currency = "USD"
    if re.search(r"€|EUR", text): currency = "EUR"
    elif re.search(r"₡|CRC|colones", text, re.IGNORECASE): currency = "CRC"

    return {
        "file":           pdf_path.name,
        "invoice_number": invoice_number,
        "date":           date,
        "vendor":         vendor[:60] if vendor else "",
        "total":          total,
        "currency":       currency,
        "raw_text_len":   len(text),
    }


# ── Validation ─────────────────────────────────────────────────────────────────
REQUIRED_FIELDS = ["invoice_number", "date", "vendor", "total"]

def validate(record: dict) -> tuple[bool, list]:
    """Validate extracted invoice data. Returns (is_valid, list_of_errors)."""
    errors = []
    for field in REQUIRED_FIELDS:
        if not record.get(field):
            errors.append(f"Missing: {field}")

    if record.get("total"):
        clean = record["total"].replace(",", "").replace("$", "")
        try:
            val = float(clean)
            if val <= 0:
                errors.append("Total must be > 0")
        except ValueError:
            errors.append(f"Invalid total format: {record['total']}")

    return (len(errors) == 0, errors)


# ── Duplicate Detection ────────────────────────────────────────────────────────
def detect_duplicates(records: list) -> list:
    """Mark duplicate invoice numbers."""
    seen = {}
    for rec in records:
        inv = rec.get("invoice_number", "").strip().upper()
        if not inv:
            rec["duplicate"] = False
            continue
        if inv in seen:
            rec["duplicate"] = True
            seen[inv]["duplicate"] = True
            log.warning(f"Duplicate invoice number: {inv} in {rec['file']}")
        else:
            seen[inv] = rec
            rec["duplicate"] = False
    return records


# ── Excel Export ───────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="FF6B35")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="1E1E1E")
VALID_FILL   = PatternFill("solid", fgColor="0D3320")
INVALID_FILL = PatternFill("solid", fgColor="3B0D0D")
DUP_FILL     = PatternFill("solid", fgColor="3B2A00")

COLUMNS = ["file", "invoice_number", "date", "vendor", "total", "currency",
           "valid", "duplicate", "errors"]

def export_excel(records: list) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    ws.sheet_view.showGridLines = False

    # Header
    for ci, col in enumerate(COLUMNS, 1):
        cell = ws.cell(1, ci, col.replace("_", " ").title())
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = max(len(col) + 6, 16)

    # Rows
    for ri, rec in enumerate(records, 2):
        row_fill = INVALID_FILL if not rec.get("valid") else (DUP_FILL if rec.get("duplicate") else (ALT_FILL if ri % 2 == 0 else None))
        for ci, col in enumerate(COLUMNS, 1):
            val = rec.get(col, "")
            if isinstance(val, list): val = "; ".join(val)
            elif isinstance(val, bool): val = "Yes" if val else "No"
            cell = ws.cell(ri, ci, val)
            cell.font = Font(color="F0EEE8", size=10)
            cell.alignment = Alignment(horizontal="center")
            if row_fill: cell.fill = row_fill

    path = OUTPUT_DIR / f"invoices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(path)
    log.info(f"Excel saved: {path.name}")
    return path


# ── CSV Export ─────────────────────────────────────────────────────────────────
def export_csv(records: list) -> Path:
    df = pd.DataFrame(records)[COLUMNS]
    df["errors"] = df["errors"].apply(lambda x: "; ".join(x) if isinstance(x, list) else x)
    path = OUTPUT_DIR / f"invoices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    df.to_csv(path, index=False)
    log.info(f"CSV saved: {path.name}")
    return path


# ── PDF Report ─────────────────────────────────────────────────────────────────
class ReportPDF(FPDF):
    def header(self):
        self.set_fill_color(13, 13, 13)
        self.rect(0, 0, 210, 18, "F")
        self.set_font("Helvetica", "B", 11)
        self.set_text_color(255, 107, 53)
        self.cell(0, 18, "Invoice Processing Report", align="L", ln=True)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "", 8)
        self.set_text_color(136, 136, 136)
        self.cell(0, 10, f"Page {self.page_no()} | Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", align="C")


def export_pdf_report(records: list, stats: dict) -> Path:
    pdf = ReportPDF()
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()
    pdf.set_fill_color(13, 13, 13)
    pdf.rect(0, 0, 210, 297, "F")

    # Title
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_text_color(255, 107, 53)
    pdf.ln(8)
    pdf.cell(0, 10, "Invoice Processing Report", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(136, 136, 136)
    pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Total PDFs: {stats['total']}", ln=True)
    pdf.ln(6)

    # Summary stats
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(240, 238, 232)
    pdf.cell(0, 8, "Summary", ln=True)
    pdf.ln(2)

    stat_items = [
        ("Total invoices",  stats["total"],      (255,107,53)),
        ("Valid",           stats["valid"],       (16,185,129)),
        ("Invalid",         stats["invalid"],     (239,68,68)),
        ("Duplicates",      stats["duplicates"],  (245,158,11)),
        ("Total amount",    f"${stats['total_amount']:,.2f}", (59,130,246)),
    ]

    for label, val, color in stat_items:
        pdf.set_fill_color(20, 20, 20)
        pdf.set_text_color(*color)
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(80, 8, label, fill=True)
        pdf.set_text_color(240, 238, 232)
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(0, 8, str(val), fill=True, ln=True)
        pdf.ln(1)

    pdf.ln(8)

    # Invoice table
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(240, 238, 232)
    pdf.cell(0, 8, "Invoice Details", ln=True)
    pdf.ln(2)

    headers = ["File", "Invoice #", "Date", "Vendor", "Total", "Status"]
    widths  = [35, 25, 22, 55, 22, 22]

    pdf.set_fill_color(255, 107, 53)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    for h, w in zip(headers, widths):
        pdf.cell(w, 7, h, fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    for i, rec in enumerate(records):
        if rec.get("duplicate"):
            pdf.set_fill_color(59, 42, 0)
        elif not rec.get("valid"):
            pdf.set_fill_color(59, 13, 13)
        elif i % 2 == 0:
            pdf.set_fill_color(20, 20, 20)
        else:
            pdf.set_fill_color(26, 26, 26)

        status = "⚠ DUP" if rec.get("duplicate") else ("✓ OK" if rec.get("valid") else "✗ ERR")
        pdf.set_text_color(240, 238, 232)
        row = [
            rec.get("file", "")[:18],
            rec.get("invoice_number", "")[:12],
            rec.get("date", "")[:10],
            rec.get("vendor", "")[:28],
            f"${rec.get('total','0')}",
            status,
        ]
        for val, w in zip(row, widths):
            pdf.cell(w, 6, str(val), fill=True)
        pdf.ln()

    path = OUTPUT_DIR / f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf.output(str(path))
    log.info(f"PDF report saved: {path.name}")
    return path


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    log.info("=== Invoice Bot started ===")

    pdf_files = list(INPUT_DIR.glob("*.pdf"))
    if not pdf_files:
        log.warning(f"No PDF files found in '{INPUT_DIR}/'. Add invoice PDFs and run again.")
        return

    log.info(f"Found {len(pdf_files)} PDF(s) to process.")
    records = []

    for pdf_path in pdf_files:
        log.info(f"Processing: {pdf_path.name}")
        try:
            data = extract_invoice_data(pdf_path)
            is_valid, errors = validate(data)
            data["valid"]  = is_valid
            data["errors"] = errors
            records.append(data)
            log.info(f"  → {'VALID' if is_valid else 'INVALID'} | Invoice: {data['invoice_number']} | Total: {data['total']}")
        except Exception as e:
            log.error(f"  → Failed to process {pdf_path.name}: {e}")
            records.append({
                "file": pdf_path.name, "invoice_number": "", "date": "",
                "vendor": "", "total": "", "currency": "", "valid": False,
                "duplicate": False, "errors": [str(e)], "raw_text_len": 0,
            })

    records = detect_duplicates(records)

    # Stats
    total_amount = 0
    for rec in records:
        try:
            val = float(rec.get("total", "0").replace(",", "").replace("$", ""))
            total_amount += val
        except: pass

    stats = {
        "total":        len(records),
        "valid":        sum(1 for r in records if r.get("valid")),
        "invalid":      sum(1 for r in records if not r.get("valid")),
        "duplicates":   sum(1 for r in records if r.get("duplicate")),
        "total_amount": total_amount,
    }

    log.info(f"Results: {stats['valid']} valid | {stats['invalid']} invalid | {stats['duplicates']} duplicates")

    export_excel(records)
    export_csv(records)
    export_pdf_report(records, stats)

    log.info("=== Invoice Bot completed. Check output/ folder. ===")


if __name__ == "__main__":
    main()
